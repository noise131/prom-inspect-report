# encoding: utf-8

from openpyxl.styles import Alignment
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime


def auto_column_width(sheet, start_row: int = 1, start_column: int | str = 1, allow_row_null: int = 10, allow_column_null: int = 10, extra_width: int = 2):
    column = start_column
    column_null = 0
    # 循环列
    while True:
        row = start_row
        row_null = 0
        max = 0
        # 循环行
        while True:
            cell_value = sheet.cell(row=row, column=column).value
            if cell_value:
                # print('{}{}'.format(get_column_letter(column), row), cell_value)
                width = len(str(cell_value).encode('gbk'))
                max = width if width > max else max
                row_null = 0
            else:
                row_null += 1
            column_char = get_column_letter(column)
            sheet['{}{}'.format(column_char, row)].alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            if row_null >= allow_row_null:
                if row <= allow_row_null + 1:
                    column_null += 1
                
                if max != 0:
                    sheet.column_dimensions[column_char].width = int(max + extra_width)
                break
        column += 1
        if column_null >= allow_column_null:
            break
